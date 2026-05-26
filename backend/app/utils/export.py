import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def _header_style(ws, row_num: int, fill_hex: str):
    fill = PatternFill("solid", fgColor=fill_hex)
    for cell in ws[row_num]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def generate_excel_export(matched: list[dict], unmatched: list[dict], summary: dict) -> bytes:
    wb = Workbook()

    # ── Summary sheet ──────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append(["Metric", "Value"])
    _header_style(ws_sum, 1, "1565C0")
    for k, v in summary.items():
        ws_sum.append([k.replace("_", " ").title(), v])
    ws_sum.column_dimensions["A"].width = 32
    ws_sum.column_dimensions["B"].width = 20

    # ── Matched Records sheet ──────────────────────────────────────────────────
    ws_matched = wb.create_sheet("Matched Records")
    matched_headers = ["Invoice Number", "GSTIN", "Vendor Name", "Taxable Amount", "IGST", "CGST", "SGST", "Match Status"]
    ws_matched.append(matched_headers)
    _header_style(ws_matched, 1, "2E7D32")
    for r in matched:
        ws_matched.append([
            r.get("invoice_number", ""),
            r.get("gstin", ""),
            r.get("vendor_name", ""),
            r.get("taxable_amount", 0),
            r.get("igst", 0),
            r.get("cgst", 0),
            r.get("sgst", 0),
            r.get("match_status", ""),
        ])
    for col in ws_matched.columns:
        ws_matched.column_dimensions[col[0].column_letter].width = 18

    # ── Unmatched Records sheet ────────────────────────────────────────────────
    ws_unmatched = wb.create_sheet("Unmatched Records")
    unmatched_headers = ["Invoice Number", "GSTIN", "Mismatch Reason", "Missing Source", "Amount Difference"]
    ws_unmatched.append(unmatched_headers)
    _header_style(ws_unmatched, 1, "C62828")
    for r in unmatched:
        ws_unmatched.append([
            r.get("invoice_number", ""),
            r.get("gstin", ""),
            r.get("mismatch_reason", ""),
            r.get("missing_source", ""),
            r.get("amount_difference", 0),
        ])
    for col in ws_unmatched.columns:
        ws_unmatched.column_dimensions[col[0].column_letter].width = 28

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
